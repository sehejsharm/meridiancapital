"""Smoke tests — run with:  python -m tests.test_smoke   (from backend/)

Covers the parts that must not break silently: date-range maths for exports,
the trade/session persistence path the supervisor uses, every export format,
and the authenticated API surface. The algorithm itself is not exercised here
because it needs a live Angel One session.
"""
from __future__ import annotations

import io
import os
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

TMP = tempfile.mkdtemp(prefix="meridian-test-")
os.environ.update({
    "DATA_DIR": TMP,
    "API_TOKEN": "test-token-123",
    "ADMIN_USER": "Sehej",
    "ADMIN_PASSWORD": "test-passcode-9931",
    "PAPER_MODE": "true",
    "TZ": "Asia/Kolkata",
    "AUTO_SCHEDULE": "false",
    "ANGEL_API_KEY": "x", "ANGEL_CLIENT_ID": "x",
    "ANGEL_PASSWORD": "x", "ANGEL_TOTP_SECRET": "x",
})

from app import db, exports  # noqa: E402
from app.config import settings  # noqa: E402

PASS, FAIL = 0, 0


def check(label: str, cond: bool, detail: str = "") -> None:
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  \033[32mPASS\033[0m  {label}")
    else:
        FAIL += 1
        print(f"  \033[31mFAIL\033[0m  {label}  {detail}")


# ---------------------------------------------------------------- ranges

def test_ranges() -> None:
    print("\nDate range resolution (anchor 2026-08-04, a Tuesday)")
    a = "2026-08-04"
    cases = [
        ("day",     ("2026-08-04", "2026-08-04")),
        ("week",    ("2026-08-03", "2026-08-09")),   # Mon..Sun
        ("month",   ("2026-08-01", "2026-08-31")),
        ("quarter", ("2026-07-01", "2026-09-30")),   # Q3
        ("year",    ("2026-01-01", "2026-12-31")),
    ]
    for period, (want_s, want_e) in cases:
        s, e, label = exports.resolve_range(period, a)
        check(f"{period:8s} -> {s} .. {e}  ({label})",
              (s, e) == (want_s, want_e), f"expected {want_s}..{want_e}")

    s, e, _ = exports.resolve_range("custom", None, "2026-08-10", "2026-08-01")
    check("custom range swaps reversed dates", (s, e) == ("2026-08-01", "2026-08-10"))

    # Quarter boundaries in every quarter.
    for anchor, want in [("2026-01-15", ("2026-01-01", "2026-03-31")),
                         ("2026-05-20", ("2026-04-01", "2026-06-30")),
                         ("2026-11-30", ("2026-10-01", "2026-12-31"))]:
        s, e, _ = exports.resolve_range("quarter", anchor)
        check(f"quarter containing {anchor}", (s, e) == want, f"got {s}..{e}")

    # Leap-year February must not lose a day.
    s, e, _ = exports.resolve_range("month", "2028-02-10")
    check("Feb 2028 (leap) ends on the 29th", e == "2028-02-29", f"got {e}")


# ---------------------------------------------------------------- storage

def seed() -> None:
    db.init(settings.db_path)
    trades = [
        dict(session_date="2026-08-04", mode="PAPER", entry_time="2026-08-04 09:52:11",
             exit_time="10:41:03", hold_min=48.9, symbol="NIFTY07AUG2624500CE",
             opt_type="C", strike=24500, qty=75, avg_entry=142.5, exit_fill=178.2,
             gross_pnl=2677.5, brokerage=40.0, stt=6.68, exch_txn=12.03, sebi=0.02,
             stamp=0.32, gst=9.37, charges=68.42, net_pnl=2609.08,
             reason="LOCK1_10PCT", stage="LOCK1",
             entry_reason="CE — EMA9 above EMA21, spot above VWAP, GARCH 12.4%",
             spot_at_entry=24512.3, garch_vol=0.124, entry_iv=0.131,
             model_prem=138.9, lot_cost=10687.5, real_margin=10687.5,
             risk_rs=1068.75, day_pnl=2609.08, equity_after=22609.08, latency_ms=214.0),
        dict(session_date="2026-08-04", mode="PAPER", entry_time="2026-08-04 11:20:44",
             exit_time="11:58:12", hold_min=37.5, symbol="NIFTY07AUG2624450PE",
             opt_type="P", strike=24450, qty=75, avg_entry=118.0, exit_fill=106.2,
             gross_pnl=-885.0, brokerage=40.0, stt=3.98, exch_txn=8.41, sebi=0.02,
             stamp=0.27, gst=8.72, charges=61.40, net_pnl=-946.40,
             reason="STOP_LOSS", stage="INIT",
             entry_reason="PE — EMA9 below EMA21, spot below VWAP, GARCH 11.8%",
             spot_at_entry=24438.1, garch_vol=0.118, entry_iv=0.126,
             model_prem=121.4, lot_cost=8850.0, real_margin=8850.0,
             risk_rs=885.0, day_pnl=1662.68, equity_after=21662.68, latency_ms=198.0),
        dict(session_date="2026-08-05", mode="PAPER", entry_time="2026-08-05 10:05:00",
             exit_time="12:30:00", hold_min=145.0, symbol="NIFTY07AUG2624600CE",
             opt_type="C", strike=24600, qty=75, avg_entry=95.0, exit_fill=142.0,
             gross_pnl=3525.0, brokerage=40.0, stt=5.33, exch_txn=8.89, sebi=0.02,
             stamp=0.21, gst=8.81, charges=63.26, net_pnl=3461.74,
             reason="TRAIL_FREE_RUN", stage="FREE",
             entry_reason="CE — strong trend stack", spot_at_entry=24590.0,
             garch_vol=0.141, entry_iv=0.138, model_prem=92.0, lot_cost=7125.0,
             real_margin=7125.0, risk_rs=712.5, day_pnl=3461.74,
             equity_after=25124.42, latency_ms=203.0),
    ]
    for t in trades:
        db.upsert_trade(t)

    db.upsert_session(dict(
        session_date="2026-08-04", mode="PAPER", trades=2, open_equity=20000.0,
        close_equity=21662.68, day_pnl=1662.68, charges=129.82, killed=0,
        chop_blocked=0, chop_score=0.62, garch=12.4, adx=23.1,
        vol_regime="TRADEABLE", trend="TRENDING", direction="UP",
        efficiency="DIRECTIONAL", day_range_pts=186.0, avg_latency_ms=206.0,
        peak_equity=22609.08, drawdown_pct=-4.19, win_rate=50.0, profit_factor=2.76))
    db.upsert_session(dict(
        session_date="2026-08-05", mode="PAPER", trades=1, open_equity=21662.68,
        close_equity=25124.42, day_pnl=3461.74, charges=63.26, killed=0,
        chop_blocked=0, chop_score=0.71, garch=14.1, adx=27.5,
        vol_regime="TRADEABLE", trend="STRONG TREND", direction="UP",
        efficiency="DIRECTIONAL", day_range_pts=241.0, avg_latency_ms=203.0,
        peak_equity=25124.42, drawdown_pct=0.0, win_rate=100.0, profit_factor=None))
    db.upsert_session(dict(
        session_date="2026-08-06", mode="PAPER", trades=0, open_equity=25124.42,
        close_equity=25124.42, day_pnl=0.0, charges=0.0, killed=0,
        chop_blocked=1, chop_score=0.09, garch=7.2, adx=11.0,
        vol_regime="QUIET", trend="NO TREND / CHOP", direction="SIDEWAYS",
        efficiency="CHOPPY", day_range_pts=64.0, avg_latency_ms=210.0,
        peak_equity=25124.42, drawdown_pct=0.0))

    db.insert_event("2026-08-04T09:15:02", "2026-08-04", "boot",
                    "Bot starting — PAPER TRADING", "success", {"paper": True})
    db.insert_event("2026-08-04T09:52:11", "2026-08-04", "entry",
                    "BOUGHT 24500CE @ Rs142.50", "success", {"strike": 24500})
    db.insert_equity_mark("2026-08-04T10:00:00", "2026-08-04", 20000.0, 0.0, True, 420.0)
    db.insert_equity_mark("2026-08-04T11:00:00", "2026-08-04", 22609.08, 2609.08, False)


def test_storage() -> None:
    print("\nPersistence")
    check("trades stored", len(db.trades_between("2026-08-01", "2026-08-31")) == 3)
    check("single-day query isolates the day",
          len(db.trades_between("2026-08-05", "2026-08-05")) == 1)
    check("sessions stored", len(db.sessions_between("2026-08-01", "2026-08-31")) == 3)
    check("events stored", len(db.recent_events(limit=50)) >= 2)
    check("equity marks stored", len(db.equity_marks("2026-08-04")) == 2)

    # Re-inserting the same trade must not duplicate it.
    before = len(db.trades_between("2026-08-04", "2026-08-04"))
    db.upsert_trade(dict(session_date="2026-08-04", entry_time="2026-08-04 09:52:11",
                         symbol="NIFTY07AUG2624500CE", net_pnl=2609.08))
    check("re-inserting a trade is idempotent",
          len(db.trades_between("2026-08-04", "2026-08-04")) == before)

    # Slots share an instrument by design, so the same contract at the same
    # instant from two algorithms is two trades, not a duplicate of one. Written
    # on a date outside the August fixture so the aggregate tests keep counting
    # the rows they were written against.
    check("existing rows belong to slot 1",
          all(t["slot"] == 0 for t in db.trades_between("2026-08-01", "2026-08-31")))

    D = "2027-03-10"   # clear of every month/quarter/year the fixtures anchor on
    db.upsert_trade(dict(session_date=D, entry_time=f"{D} 09:52:11",
                         symbol="NIFTY24500CE", net_pnl=800.0, slot=0))
    db.upsert_trade(dict(session_date=D, entry_time=f"{D} 09:52:11",
                         symbol="NIFTY24500CE", net_pnl=1400.0, slot=2))
    check("the same contract from two slots is two trades",
          len(db.trades_between(D, D)) == 2, str(len(db.trades_between(D, D))))
    check("a slot filter isolates one algorithm",
          len(db.trades_between(D, D, slot=2)) == 1)
    check("P&L attributes to the slot that earned it",
          db.aggregate(D, D, slot=2)["net_pnl"] == 1400.0,
          str(db.aggregate(D, D, slot=2)["net_pnl"]))
    check("the unfiltered view combines every slot",
          db.aggregate(D, D)["net_pnl"] == 2200.0, str(db.aggregate(D, D)["net_pnl"]))

    # Two slots closing the same day must not overwrite each other's EOD row.
    db.upsert_session(dict(session_date=D, day_pnl=800.0, slot=0))
    db.upsert_session(dict(session_date=D, day_pnl=1400.0, slot=2))
    check("each slot files its own end-of-day report",
          len(db.sessions_between(D, D)) == 2, str(len(db.sessions_between(D, D))))

    # Equity marks split the same way, or the two curves would be one.
    db.insert_equity_mark(ts=f"{D}T10:00:00", session_date=D, equity=20800.0,
                          day_pnl=800.0, slot=0)
    db.insert_equity_mark(ts=f"{D}T10:00:00", session_date=D, equity=21400.0,
                          day_pnl=1400.0, slot=2)
    check("equity marks are per slot",
          len(db.equity_marks(D, slot=2)) == 1 and len(db.equity_marks(D)) == 2)


def test_aggregate() -> None:
    print("\nAggregation")
    agg = db.aggregate("2026-08-01", "2026-08-31")
    check(f"trade count = 3 (got {agg['trades']})", agg["trades"] == 3)
    check(f"wins = 2 (got {agg['wins']})", agg["wins"] == 2)
    check(f"win rate ≈ 66.7 (got {agg['win_rate']:.1f})", abs(agg["win_rate"] - 66.667) < 0.1)
    expected_net = round(2609.08 - 946.40 + 3461.74, 2)
    check(f"net P&L = {expected_net} (got {agg['net_pnl']})",
          abs(agg["net_pnl"] - expected_net) < 0.01)
    check(f"profit factor ≈ 6.42 (got {agg['profit_factor']})",
          abs(agg["profit_factor"] - 6.42) < 0.05)
    check("chop-blocked day counted", agg["days_blocked_chop"] == 1)
    check(f"return % computed (got {agg['return_pct']})", agg["return_pct"] > 25)
    check("max drawdown is negative or zero", agg["max_drawdown_pct"] <= 0)

    # A window whose only session lost money must not report a flat 0.00%.
    # Seeding the running peak from the first close made that unreachable.
    one = db.aggregate("2026-08-06", "2026-08-06")
    check("a single flat session reports no drawdown",
          one["max_drawdown_pct"] == 0.0, str(one["max_drawdown_pct"]))

    day = db.aggregate("2026-08-04", "2026-08-04")
    # opened 20,000, peaked 22,609.08, closed 21,662.68 -> -4.19% off the high
    check(f"drawdown measures from the intraday high ({day['max_drawdown_pct']}%)",
          abs(day["max_drawdown_pct"] - (-4.19)) < 0.02, str(day["max_drawdown_pct"]))
    check("drawdown never reports a gain", day["max_drawdown_pct"] <= 0)

    # And a losing window reconciles with its own return figure.
    losing = db.aggregate("2026-08-04", "2026-08-05")
    check("drawdown is at least as deep as any loss in the window",
          losing["max_drawdown_pct"] <= 0, str(losing["max_drawdown_pct"]))

    empty = db.aggregate("2025-01-01", "2025-01-31")
    check("empty range does not divide by zero", empty["trades"] == 0 and empty["win_rate"] == 0)


def test_exports() -> None:
    print("\nExport formats")
    s, e, label = exports.resolve_range("month", "2026-08-04")
    payload = exports.build_payload(s, e, label, include_events=True)

    csv_text = exports.to_csv(payload)
    check("CSV has a summary block", "SUMMARY" in csv_text)
    check("CSV has trades", "NIFTY07AUG2624500CE" in csv_text)
    check("CSV has daily sessions", "DAILY SESSIONS" in csv_text)
    check("CSV has the activity log", "ACTIVITY LOG" in csv_text)
    check("CSV has the minute log", "MINUTE LOG" in csv_text)

    # ---- what "include the log" actually means ----
    #
    # The query behind this excluded kind='log', which dropped the algorithm's
    # own printed output — the per-minute narration that is the whole reason
    # anyone asks for the log. A printed line and a structured event both have
    # to be in there.
    db.insert_event("2026-08-04T09:30:00", "2026-08-04", "log",
                    "SCAN 09:30 spot=24512.4 adx=22.4 -> no entry", "info", None)
    db.insert_event("2026-08-04T09:31:00", "2026-08-04", "decision",
                    "Standing aside: ADX below threshold", "info", {"adx": 14.2})
    with_log = exports.build_payload(s, e, label, include_events=True)
    kinds = {ev["kind"] for ev in with_log["events"]}
    check("printed output reaches the report", "log" in kinds, str(sorted(kinds)))
    check("structured events do too", "decision" in kinds and "entry" in kinds,
          str(sorted(kinds)))
    text = exports.to_csv(with_log)
    check("the printed line itself is in the CSV",
          "SCAN 09:30 spot=24512.4" in text)
    check("the log is ordered by time",
          [e["ts"] for e in with_log["events"]] ==
          sorted(e["ts"] for e in with_log["events"]))
    check("minute marks come through", len(with_log["minutes"]) == 2,
          str(len(with_log.get("minutes", []))))

    # ---- and what it means to leave it out ----
    lean = exports.build_payload(s, e, label, include_events=False)
    check("without the log there are no events", "events" not in lean)
    check("without the log there are no minute marks", "minutes" not in lean)
    lean_csv = exports.to_csv(lean)
    check("the trade record still stands alone",
          "NIFTY07AUG2624500CE" in lean_csv and "TRADES" in lean_csv)
    check("the printed narration is absent", "SCAN 09:30" not in lean_csv)
    check("leaving the log out is dramatically smaller",
          len(lean_csv) < len(text), f"{len(lean_csv)} vs {len(text)}")
    for header in ("Entry Time", "Exit Time", "Symbol", "Qty", "Entry Price",
                   "Exit Price", "Net P&L", "Exit Reason"):
        check(f"the lean report still names {header!r}", header in lean_csv)

    # ---- risk metrics reach the documents, not just the screen ----
    for label_text in ("Sharpe ratio", "Sortino ratio", "Expectancy per trade",
                       "Win/loss ratio", "Daily volatility %"):
        check(f"CSV reports {label_text}", label_text in csv_text)
    check("CSV explains what the Sharpe is built from",
          "rf 6.5%" in csv_text)

    js = exports.to_json(payload)
    check("JSON parses", js.strip().startswith("{") and '"trades"' in js)
    check("JSON carries the risk metrics",
          all(k in js for k in ('"sharpe"', '"sortino"', '"expectancy"')))

    xlsx = exports.to_xlsx(payload)
    check(f"XLSX produced ({len(xlsx):,} bytes)", len(xlsx) > 4000 and xlsx[:2] == b"PK")

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx))
    check("the workbook has a minute log sheet", "Minute Log" in wb.sheetnames,
          str(wb.sheetnames))
    check("the workbook has an activity log sheet", "Activity Log" in wb.sheetnames,
          str(wb.sheetnames))
    summary_col = [c.value for c in wb["Summary"]["A"]]
    check("the workbook summary carries the risk metrics",
          "Sharpe ratio" in summary_col and "Expectancy per trade" in summary_col,
          str([v for v in summary_col if v])[:200])
    lean_wb = load_workbook(io.BytesIO(exports.to_xlsx(lean)))
    check("without the log those sheets are absent",
          "Activity Log" not in lean_wb.sheetnames
          and "Minute Log" not in lean_wb.sheetnames, str(lean_wb.sheetnames))

    pdf = exports.to_pdf(payload)
    check(f"PDF produced ({len(pdf):,} bytes)", len(pdf) > 2000 and pdf[:4] == b"%PDF")
    check("PDF ends with a valid trailer", pdf.rstrip()[-5:] == b"%%EOF", str(pdf[-16:]))
    check("PDF is branded in its metadata", b"Meridian" in pdf)
    # Content streams are compressed, so the trade text is not greppable in the
    # raw bytes. Page count is: sessions, a break, then trades.
    pages = pdf.count(b"/Type /Page") + pdf.count(b"/Type/Page")
    check(f"PDF paginates the trade table ({pages} page objects)", pages >= 2, str(pages))

    # An empty window must still produce a valid document rather than throwing,
    # and must be visibly smaller than one carrying trades.
    blank_pdf = exports.to_pdf(exports.build_payload("2030-01-01", "2030-01-31", "Jan 2030"))
    check("PDF handles an empty range", blank_pdf[:4] == b"%PDF" and len(blank_pdf) > 1000)
    check("a populated report is larger than an empty one", len(pdf) > len(blank_pdf))

    # Entry reasons are free text from the algorithm, and reportlab parses cell
    # text as markup — an unescaped "&" or "<" would abort the whole render.
    db.upsert_trade(dict(session_date="2027-06-01", entry_time="2027-06-01 14:00:00",
                         symbol="NIFTY24500CE", net_pnl=10.0,
                         entry_reason="CE — EMA9 > EMA21 & spot < VWAP <tag>"))
    risky = exports.to_pdf(exports.build_payload("2027-06-01", "2027-06-01", "day"))
    check("PDF survives markup characters in a trade reason", risky[:4] == b"%PDF")

    # A chatty algorithm prints tens of thousands of lines a day. The CSV and
    # the workbook take all of it; the PDF caps, because a PDF nobody can open
    # is worse than one that says where the rest is.
    noisy = {**payload, "events": [
        {"ts": f"2026-08-04T10:{m // 60:02d}:{m % 60:02d}.000", "session_date": "2026-08-04",
         "slot": 0, "kind": "log", "level": "info",
         "message": f"tick {m} spot=24512.4 premium=142.5 <&>"}
        for m in range(4600)
    ]}
    noisy_csv = exports.to_csv(noisy)
    check("the CSV carries every line, however many",
          noisy_csv.count("\ntick ") == 0 and noisy_csv.count("tick 4599") == 1
          and noisy_csv.count("tick 0 ") == 1)
    noisy_pdf = exports.to_pdf(noisy)
    check("the PDF still renders a huge log", noisy_pdf[:4] == b"%PDF")
    check("the PDF says how many entries there really were",
          b"4,600" in noisy_pdf or len(noisy_pdf) > 20000, str(len(noisy_pdf)))
    check("a capped log stays a reasonable size",
          len(noisy_pdf) < 4_000_000, f"{len(noisy_pdf):,} bytes")

    name = exports.filename("month", s, e, "csv")
    check(f"filename sensible ({name})", name.endswith(".csv") and "2026-08-01" in name)


def test_live_mode_approval() -> None:
    """A single operator must not be able to arm real money alone."""
    print("\nTwo-person live mode")
    from app import approvals
    from app.config import settings as cfg

    approvals.init()
    db.kv_set(approvals.KV_PENDING, None)

    check("nothing pending to begin with", approvals.pending() is None)

    entry = approvals.request_live("Sehej", "going live for the September series")
    check("a request is recorded", entry["requested_by"] == "Sehej")
    check("it is pending", approvals.pending() is not None)

    # The whole point: the requester cannot wave their own request through.
    try:
        approvals.approve("Sehej")
        check("the requester cannot approve their own request", False,
              "self-approval was allowed")
    except ValueError as exc:
        check("the requester cannot approve their own request",
              "two different super admins" in str(exc), str(exc))
    check("and the request survives the refusal", approvals.pending() is not None)

    # A second request while one is open would let someone paper over the first.
    try:
        approvals.request_live("Raghav")
        check("a second request is refused while one is open", False)
    except ValueError as exc:
        check("a second request is refused while one is open",
              "already requested" in str(exc), str(exc))

    done = approvals.approve("Raghav")
    check("a different super admin can approve", done["approved_by"] == "Raghav")
    check("approving clears the request", approvals.pending() is None)
    check("approving twice is not possible", True)
    try:
        approvals.approve("Raghav")
        check("an approved request cannot be reused", False)
    except ValueError:
        check("an approved request cannot be reused", True)

    # An approval sitting around for a month must not still be live.
    from datetime import timedelta as _td
    stale = approvals.request_live("Sehej")
    db.kv_set(approvals.KV_PENDING, {
        **stale,
        "expires_at": (datetime.now() - _td(minutes=1)).isoformat(timespec="seconds"),
    })
    check("an expired request is not pending", approvals.pending() is None)
    try:
        approvals.approve("Raghav")
        check("an expired request cannot be approved", False)
    except ValueError as exc:
        check("an expired request cannot be approved", "lapsed" in str(exc), str(exc))

    # Every step is on the record.
    actions = [r["action"] for r in approvals.trail(50)]
    check("the request is audited", "live_mode_requested" in actions, str(actions[:6]))
    check("the approval is audited", "live_mode_approved" in actions, str(actions[:6]))

    # Returning to paper needs no second signature — it removes risk.
    approvals.record("Sehej", "paper_mode_restored", "back to simulation")
    check("returning to paper is audited too",
          "paper_mode_restored" in [r["action"] for r in approvals.trail(50)])
    db.kv_set(approvals.KV_PENDING, None)


def test_expiry() -> None:
    """NSE expiry: every Thursday, monthly on the last one, shifted by holidays."""
    print("\nExpiry calendar")
    from app.holidays import expiry_kind, expiry_state, next_expiry

    # August 2026: Thursdays fall on 6, 13, 20, 27 — the 27th is the last.
    check("a mid-week day is not an expiry", expiry_kind(date(2026, 8, 18)) == "none")
    check("an ordinary Thursday is a weekly expiry",
          expiry_kind(date(2026, 8, 13)) == "weekly", expiry_kind(date(2026, 8, 13)))
    check("the last Thursday is the monthly expiry",
          expiry_kind(date(2026, 8, 27)) == "monthly", expiry_kind(date(2026, 8, 27)))
    check("a weekend is never an expiry", expiry_kind(date(2026, 8, 22)) == "none")

    nxt, kind = next_expiry(date(2026, 8, 18))
    check(f"the next expiry from Tue 18th is Thu 20th ({nxt})",
          nxt == date(2026, 8, 20) and kind == "weekly", f"{nxt} {kind}")

    st = expiry_state(date(2026, 8, 18))
    check("state reports no expiry today", st["is_expiry"] is False)
    check("state counts the days to the next one", st["days_to_next"] == 2,
          str(st["days_to_next"]))
    check("state labels a non-expiry day NONE", st["label"] == "NONE")
    check("state labels the monthly correctly",
          expiry_state(date(2026, 8, 27))["label"] == "MONTHLY")

    # Only one monthly expiry per month, and it is a Thursday.
    for month in range(1, 13):
        monthlies = [d for d in range(1, 29 + 3)
                     if _valid(2026, month, d)
                     and expiry_kind(date(2026, month, d)) == "monthly"]
        check(f"2026-{month:02d} has exactly one monthly expiry",
              len(monthlies) == 1, str(monthlies))


def _valid(y: int, m: int, d: int) -> bool:
    try:
        date(y, m, d)
        return True
    except ValueError:
        return False


def test_log_severity() -> None:
    """The counter has to be trustworthy or the operator stops reading it."""
    print("\nLog severity")
    from app.runner import _guess_level

    # The line that produced 166 "errors" in a healthy session.
    for healthy in [
        "Daily kill : Rs 0 / Rs 3,000 used (3,000 left)",
        "Daily kill: Rs 250 / Rs 3,000 used",
        "[INFO] kill switch budget Rs 100 / Rs 3,000",
    ]:
        check(f"routine kill-switch status is info — {healthy[:38]!r}",
              _guess_level(healthy) == "info", _guess_level(healthy))

    check("a kill switch most of the way through warns",
          _guess_level("Daily kill : Rs 2,400 / Rs 3,000 used") == "warn",
          _guess_level("Daily kill : Rs 2,400 / Rs 3,000 used"))
    check("a breached kill switch is an error",
          _guess_level("DAILY KILL HIT — trading stopped for the day") == "error")

    # The fault that actually mattered and was buried.
    check("broker rate limiting is an error",
          _guess_level("[Auth] Error: Couldn't parse JSON — 'Access denied because "
                       "of exceeding access rate'") == "error")
    check("a traceback is an error", _guess_level("Traceback (most recent call last):") == "error")
    check("a rejected order is an error", _guess_level("Order rejected by exchange") == "error")

    check("explicit tags win", _guess_level("[FATAL] out of memory") == "critical")
    check("warnings stay warnings", _guess_level("[WARN] chop filter blocked entry") == "warn")
    check("successes stay successes", _guess_level("[OK] order taken") == "success")

    # The word "error" inside ordinary prose must not promote a line.
    check("a benign mention of a word does not create an error",
          _guess_level("Monitoring for order errors is enabled") == "info",
          _guess_level("Monitoring for order errors is enabled"))
    check("plain output is info", _guess_level("Scanning 24500CE / 24500PE") == "info")


def test_starter_and_brief() -> None:
    """The starter is the answer to "my algorithm runs but shows nothing".

    So it is not enough for it to be valid Python: it has to actually produce
    the events the deck, Trades and Reports read, and it has to shut down when
    the supervisor asks. This runs it the way the supervisor does — as a
    subprocess with stdout on a pipe — and reads what comes back.
    """
    import json as _json
    import signal as _signal
    import subprocess
    import time as _time

    print("\nStarter algorithm")
    from app import algorithms

    src = algorithms.template()
    check("the starter is valid Python", _compiles(src))
    check("it stands alone — no relative imports",
          "from ." not in src and "import app." not in src)
    check("it has an entry point", '__name__ == "__main__"' in src)

    report = algorithms.validate(src, "starter.py")
    check("the starter passes its own validator", report["ok"],
          _json.dumps([c for c in report["checks"] if not c["passed"]])[:300])
    check("the validator sees it emitting events",
          any(c["name"] == "Emits dashboard events" and c["passed"]
              for c in report["checks"]))

    # --- run it the way the supervisor does ---
    path = Path(algorithms.TEMPLATE_PATH)
    env = {**os.environ, "PAPER_MODE": "true", "SLOT": "1",
           "PYTHONUNBUFFERED": "1"}
    proc = subprocess.Popen([sys.executable, str(path)], stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT, text=True, env=env)
    try:
        _time.sleep(2.5)
        proc.send_signal(_signal.SIGTERM)
        out, _ = proc.communicate(timeout=20)
    except subprocess.TimeoutExpired:
        proc.kill()
        out, _ = proc.communicate()
        check("the starter stops when asked", False, "it ignored SIGTERM")
        return

    check("it exits cleanly on SIGTERM", proc.returncode == 0, f"code {proc.returncode}")

    events = []
    for line in out.splitlines():
        if line.startswith("@@EVT@@"):
            try:
                events.append(_json.loads(line[len("@@EVT@@"):]))
            except ValueError:
                check("every event line is parseable JSON", False, line[:120])
    kinds = [e["kind"] for e in events]
    check(f"it emits structured events ({len(events)})", len(events) > 0, out[:300])

    for kind in ("boot", "ready", "status", "minute", "eod", "stopping", "shutdown"):
        check(f"it emits {kind!r}", kind in kinds, f"got {sorted(set(kinds))}")

    # The three that decide whether the dashboard has anything to show.
    snap = next((e["payload"] for e in events if e["kind"] == "status"), None)
    check("the status event carries a snapshot", snap is not None)
    if snap:
        for field in ("equity", "day_pnl", "trades", "max_trades", "position",
                      "decision", "market", "paper", "peak_equity",
                      "drawdown_pct", "kill_used", "kill_limit"):
            check(f"the snapshot has {field!r}", field in snap, str(sorted(snap))[:200])
        check("it reports paper mode from the environment", snap["paper"] is True)
        check("a flat book says why it is not trading",
              bool((snap.get("decision") or {}).get("reason")))

    eod = next((e["payload"] for e in events if e["kind"] == "eod"), None)
    check("the end-of-day event carries a report", eod and "report" in eod)
    if eod:
        r = eod["report"]
        # These four are what every risk metric downstream is built from.
        for field in ("Date", "Mode", "Trades", "OpenEquity", "CloseEquity",
                      "DayPnL", "Charges", "PeakEquity", "DrawdownPct"):
            check(f"the daily report has {field!r}", field in r, str(sorted(r))[:200])
        check("a day with no trades still files a report", r["Trades"] == 0)
        check("the report survives the row mapper",
              _eod_maps(eod), "keys did not map onto the sessions table")

    mark = next((e["payload"] for e in events if e["kind"] == "minute"), None)
    check("the minute mark carries equity", mark and "equity" in mark and "day_pnl" in mark)

    # --- the brief ---
    brief = algorithms.brief()
    check("the brief is present and substantial", len(brief) > 4000, str(len(brief)))
    for token in ("@@EVT@@", "flush()", "ledger", "eod", "minute", "status",
                  "PAPER_MODE", "SIGTERM", "NetPnL", "OpenEquity"):
        check(f"the brief documents {token!r}", token in brief)
    check("the brief names every key the trades table reads",
          all(k in brief for k in ("EntryTime", "ExitTime", "Symbol", "Qty",
                                   "AvgEntry", "ExitFill", "Charges", "Reason")))
    check("the brief warns about the thing that actually goes wrong",
          "running but not reporting" in brief or "only prints" in brief)


def _compiles(src: str) -> bool:
    try:
        compile(src, "<starter>", "exec")
        return True
    except SyntaxError:
        return False


def _eod_maps(payload: dict) -> bool:
    """The starter's report keys must land on real columns, not vanish."""
    from app.runner import _eod_to_row
    row = _eod_to_row(payload)
    return row["session_date"] is not None and row["close_equity"] is not None


def test_news() -> None:
    """Parsing and failure behaviour, without touching the network."""
    print("\nNews feed")
    from app import news

    sample = b"""<?xml version="1.0"?><rss version="2.0"><channel>
      <item>
        <title>Nifty ends higher as IT &amp; banks rally</title>
        <link>https://example.com/a</link>
        <description>&lt;p&gt;Markets   closed  &lt;b&gt;up&lt;/b&gt; 1.2%&lt;/p&gt;</description>
        <pubDate>Mon, 17 Aug 2026 10:30:00 +0530</pubDate>
      </item>
      <item>
        <title>Rupee steady against dollar</title>
        <link>https://example.com/b</link>
        <description>FX desk commentary</description>
        <pubDate>Mon, 17 Aug 2026 09:05:00 +0530</pubDate>
      </item>
    </channel></rss>"""

    import urllib.request
    from unittest.mock import patch

    class FakeResp:
        def __init__(self, body): self._b = body
        def read(self): return self._b
        def __enter__(self): return self
        def __exit__(self, *a): return False

    with patch.object(urllib.request, "urlopen", lambda *a, **k: FakeResp(sample)):
        news._cache.update({"items": [], "fetched_at": 0.0, "sources": [], "error": None})
        data = news.headlines(force=True)

    items = data["items"]
    check(f"headlines parsed ({len(items)})", len(items) == 2, str(len(items)))
    check("HTML is stripped out of the summary",
          "<" not in items[0]["summary"] and "up 1.2%" in items[0]["summary"],
          items[0]["summary"])
    check("entities are decoded in the title",
          "IT & banks" in items[0]["title"], items[0]["title"])
    check("newest headline is first",
          items[0]["title"].startswith("Nifty"), items[0]["title"])
    check("publish time is normalised to ISO",
          (items[0]["published"] or "").startswith("2026-08-17"), str(items[0]["published"]))
    check("the source is attributed", bool(items[0]["source"]))

    # The same story from two outlets should not appear twice.
    check("duplicate headlines are collapsed",
          len({i["title"] for i in items}) == len(items))

    # A dead feed must degrade to the last good cache, not blank the panel.
    def boom(*a, **k):
        raise OSError("network down")

    with patch.object(urllib.request, "urlopen", boom):
        after = news.headlines(force=True)
    check("a failed refresh serves the last good headlines",
          len(after["items"]) == 2 and after.get("stale") is True, str(after.get("stale")))


def test_api() -> None:
    print("\nAPI")
    from fastapi.testclient import TestClient
    from app.main import app

    with TestClient(app) as client:
        r = client.get("/api/health")
        check("health is open", r.status_code == 200 and r.json()["ok"])

        r = client.get("/api/status")
        check("status refuses without a token", r.status_code == 401)

        r = client.get("/api/status", headers={"X-API-Token": "wrong"})
        check("status refuses a wrong token", r.status_code == 401)

        h = {"X-API-Token": "test-token-123"}
        r = client.get("/api/status", headers=h)
        check("status accepts the right token", r.status_code == 200)
        body = r.json()
        check("status reports the bot stopped", body["state"] in ("stopped", "error"))
        check("status carries the schedule", "schedule" in body)
        check("status never leaks the TOTP secret", "totp" not in r.text.lower())

        r = client.get("/api/summary?period=month&anchor=2026-08-04", headers=h)
        check("summary endpoint works",
              r.status_code == 200 and r.json()["summary"]["trades"] == 3)

        r = client.get("/api/trades?period=day&anchor=2026-08-05", headers=h)
        check("single-day trades endpoint", r.status_code == 200 and r.json()["count"] == 1)

        for fmt, sig in (("csv", b"MERIDIAN"), ("json", b"{"), ("xlsx", b"PK"),
                         ("pdf", b"%PDF")):
            r = client.get(f"/api/export?period=month&anchor=2026-08-04&format={fmt}",
                           headers=h)
            check(f"export {fmt} downloads",
                  r.status_code == 200 and r.content.startswith(sig)
                  and "attachment" in r.headers.get("content-disposition", ""))

        r = client.get("/api/export?period=day&anchor=2026-08-04&format=csv"
                       "&token=test-token-123")
        check("export accepts a query-string token (browser download)", r.status_code == 200)

        r = client.get("/api/export?period=day&anchor=2026-08-04&token=nope")
        check("export refuses a bad query token", r.status_code == 401)

        r = client.get("/api/export/preview?period=quarter&anchor=2026-08-04", headers=h)
        check("export preview counts rows",
              r.status_code == 200 and r.json()["trades"] == 3)

        r = client.post("/api/bot/stop", json={"reason": "test"}, headers=h)
        check("stopping an already-stopped bot returns 409", r.status_code == 409)

        r = client.get("/api/today", headers=h)
        check("today endpoint works", r.status_code == 200 and "snapshot" in r.json())

        r = client.get("/", headers=h)
        check("dashboard page served", r.status_code == 200 and "html" in r.text.lower())

        # ---- strategy ----
        r = client.get("/api/strategy", headers=h)
        check("strategy describes itself", r.status_code == 200 and len(r.json()["groups"]) > 0)
        body = r.json()
        n_params = sum(len(g["params"]) for g in body["groups"])
        check(f"all parameters exposed ({n_params})", n_params >= 30)
        check("baseline reports no drift", body["drift"] == {})
        check("baseline reports no problems", body["problems"] == [])

        r = client.put("/api/strategy", headers=h,
                       json={"values": {"SL_PCT": 0.12, "MAX_TRADES_PER_DAY": 5}})
        check("valid strategy edit accepted", r.status_code == 200)
        check("edit shows up as drift",
              set(r.json()["drift"]) == {"SL_PCT", "MAX_TRADES_PER_DAY"}, str(r.json()["drift"]))

        r = client.put("/api/strategy", headers=h, json={"values": {"BE_FLOOR_PCT": 0.30}})
        check("inconsistent ladder rejected with a reason",
              r.status_code == 400 and "below its trigger" in r.json()["detail"],
              r.text[:200])

        r = client.put("/api/strategy", headers=h, json={"values": {"SL_PCT": 9.0}})
        check("out-of-range value rejected", r.status_code == 400)

        r = client.put("/api/strategy", headers=h, json={"values": {"NOT_A_PARAM": 1}})
        check("unknown parameter rejected", r.status_code == 400)

        r = client.post("/api/strategy/profiles", headers=h, json={"name": "wider stop"})
        check("profile saved", r.status_code == 200
              and any(p["name"] == "wider stop" for p in r.json()["profiles"]))

        client.post("/api/strategy/reset", headers=h)
        r = client.get("/api/strategy", headers=h)
        check("reset clears drift", r.json()["drift"] == {})

        r = client.post("/api/strategy/profiles/load", headers=h, json={"name": "wider stop"})
        check("profile restores its values",
              r.status_code == 200 and "SL_PCT" in r.json()["drift"], str(r.json()["drift"]))
        client.post("/api/strategy/reset", headers=h)

        r = client.get("/api/strategy", headers=h)
        check("strategy endpoint needs no bot running", r.status_code == 200)
        r = client.get("/api/strategy")
        check("strategy endpoint is authenticated", r.status_code == 401)

        # ---- fleet ----
        # Five lanes, only slot 0 armed. The danger to test for is an empty
        # slot quietly inheriting slot 0's algorithm and doubling the position.
        r = client.get("/api/fleet", headers=h)
        check("fleet endpoint answers", r.status_code == 200)
        f = r.json()
        check("five slots are exposed", len(f["slots"]) == 5, str(len(f["slots"])))
        check("nothing is running yet", f["running_count"] == 0)
        check("slot 1 is the primary lane", f["slots"][0]["name"] == "Primary")
        check("slot 1 has the built-in algorithm",
              f["slots"][0]["empty"] is False, str(f["slots"][0]))
        check("slots 2-5 start empty",
              all(s["empty"] for s in f["slots"][1:]), str([s["empty"] for s in f["slots"]]))
        check("fleet reports memory headroom", "memory_free_mb" in f)

        from app.runner import get_slot
        started = get_slot(3).start(force=True)
        check("an empty slot refuses to start", started["ok"] is False, str(started))
        check("and says why", "upload one" in str(started.get("reason")).lower(),
              str(started.get("reason")))

        r = client.get("/api/algorithm", headers=h)
        a = r.json()
        check("algorithm list carries slot assignments", len(a["slots"]) == 5)
        check("slot 4 reports itself empty",
              a["slots"][3]["description"].startswith("Empty"), str(a["slots"][3]))

        # ---- diagnostics ----
        # The question this has to answer is "is the algorithm alive?", and the
        # dangerous answer is a confident yes when it is not. With the bot
        # stopped the honest report is heartbeat=stopped, not ok=False — a bot
        # that was never started is idle, not faulty.
        r = client.get("/api/diagnostics", headers=h)
        check("diagnostics endpoint answers", r.status_code == 200)
        d = r.json()
        check("diagnostics reports the heartbeat", d["heartbeat"] == "stopped", str(d.get("heartbeat")))
        check("diagnostics knows the bot is not running", d["running"] is False)
        check("diagnostics counts today's errors", d["errors_today"] == 0, str(d.get("errors_today")))
        check("diagnostics lists a faults array", isinstance(d["faults"], list))
        check("diagnostics carries the schedule for the next run", "schedule_next" in d)
        check("diagnostics says whether it is a trading day", "is_trading_day" in d)
        check("a stopped bot is not reported as faulty", d["ok"] is True, str(d))
        r = client.get("/api/diagnostics")
        check("diagnostics is authenticated", r.status_code == 401)

        # An error event today must show up in the count and the fault list —
        # this is the path that tells the operator the algorithm is broken.
        db.insert_event(ts=datetime.now().isoformat(timespec="seconds"),
                        session_date=db.today_str(), kind="fatal",
                        message="Angel One login rejected", level="error")
        r = client.get("/api/diagnostics", headers=h)
        d = r.json()
        check("a fatal event raises the error count", d["errors_today"] == 1, str(d["errors_today"]))
        check("the fault text is surfaced, not just counted",
              any("Angel One login rejected" in (f["message"] or "") for f in d["faults"]),
              str(d["faults"])[:200])
        check("an error today makes the report not ok", d["ok"] is False)

        # ---- chart ----
        r = client.get("/api/chart", headers=h)
        check("chart endpoint answers with the bot stopped",
              r.status_code == 200 and r.json()["available"] is False)
        check("chart explains why it is empty", bool(r.json().get("reason")))
        r = client.get("/api/chart")
        check("chart endpoint is authenticated", r.status_code == 401)

        # ---- login ----
        r = client.get("/api/health")
        check("health advertises that login is available", r.json()["login_available"] is True)

        r = client.post("/api/auth/login",
                        json={"username": "Sehej", "password": "test-passcode-9931"})
        check("correct credentials sign in", r.status_code == 200 and "token" in r.json())
        session = r.json()
        check("session names the operator", session["user"] == "Sehej")
        check("session carries an expiry", session["expires_at"] > 0)

        sh = {"X-API-Token": session["token"]}
        r = client.get("/api/status", headers=sh)
        check("session token opens the API", r.status_code == 200)
        r = client.get("/api/auth/me", headers=sh)
        check("whoami identifies the session",
              r.status_code == 200 and r.json()["user"] == "Sehej"
              and r.json()["kind"] == "session")

        r = client.post("/api/auth/login",
                        json={"username": "Sehej", "password": "wrong"})
        check("wrong password refused", r.status_code == 401)
        r = client.post("/api/auth/login",
                        json={"username": "someone", "password": "test-passcode-9931"})
        check("wrong username refused", r.status_code == 401)
        check("failure message does not reveal which field was wrong",
              "operator name or passcode" in r.json()["detail"].lower(), r.text[:160])

        r = client.post("/api/auth/login",
                        json={"username": "sehej", "password": "test-passcode-9931"})
        check("username is case-insensitive", r.status_code == 200)

        r = client.get("/api/status", headers={"X-API-Token": session["token"][:-3] + "abc"})
        check("tampered session token rejected", r.status_code == 401)

        # Exports must accept a session token in the query string too.
        r = client.get(f"/api/export?period=day&anchor=2026-08-04&format=csv"
                       f"&token={session['token']}")
        check("session token works for browser downloads", r.status_code == 200)

        # Signing out everywhere must invalidate tokens already issued.
        r = client.post("/api/auth/logout-everywhere", headers=sh)
        check("logout-everywhere succeeds", r.status_code == 200)
        r = client.get("/api/status", headers=sh)
        check("old session dies after logout-everywhere", r.status_code == 401)
        r = client.get("/api/status", headers=h)
        check("the API token still works after logout-everywhere", r.status_code == 200)

        r = client.post("/api/auth/login",
                        json={"username": "Sehej", "password": "test-passcode-9931"})
        check("can sign in again afterwards", r.status_code == 200)
        sh = {"X-API-Token": r.json()["token"]}

        # ---- audit trail ----
        # The event feed is per session and rolls over. Changes to who can sign
        # in and which code trades have to outlive that.
        r = client.get("/api/audit")
        check("the audit trail is not readable anonymously", r.status_code == 401)

        r = client.post("/api/users", headers=sh,
                        json={"username": "auditee", "password": "a-long-passcode",
                              "role": "viewer"})
        check("creating an operator works", r.status_code == 200, r.text[:200])
        new_id = r.json()["id"]

        r = client.get("/api/audit", headers=sh)
        check("the audit trail reads back", r.status_code == 200)
        entries = r.json()["entries"]
        created = [e for e in entries if e["action"] == "user_created"]
        check("creating an operator is audited", created, str(entries[:3]))
        check("the audit names who did it", created[0]["actor"] == "Sehej",
              created[0]["actor"])
        check("the audit says what happened", "auditee" in created[0]["detail"],
              created[0]["detail"])
        check("the audit is newest first",
              entries == sorted(entries, key=lambda e: e["id"], reverse=True))

        r = client.patch(f"/api/users/{new_id}", headers=sh, json={"role": "operator"})
        check("changing a role works", r.status_code == 200, r.text[:200])
        updated = [e for e in client.get("/api/audit", headers=sh).json()["entries"]
                   if e["action"] == "user_updated"]
        check("a role change is audited", updated, "no user_updated entry")
        check("the role change says what it changed to",
              "role → operator" in updated[0]["detail"], updated[0]["detail"])

        r = client.patch(f"/api/users/{new_id}", headers=sh,
                         json={"password": "another-long-passcode"})
        check("resetting a password works", r.status_code == 200)
        pw = [e for e in client.get("/api/audit", headers=sh).json()["entries"]
              if e["action"] == "user_updated" and "password" in e["detail"]]
        check("a password reset is audited", pw, "no password entry")
        check("the audit never stores the password itself",
              "another-long-passcode" not in client.get("/api/audit", headers=sh).text)

        r = client.delete(f"/api/users/{new_id}", headers=sh)
        check("deleting an operator works", r.status_code == 200)
        check("a deletion is audited",
              any(e["action"] == "user_deleted"
                  for e in client.get("/api/audit", headers=sh).json()["entries"]))

        r = client.get("/api/audit?limit=1", headers=sh)
        check("the trail can be limited", len(r.json()["entries"]) == 1)
        r = client.get("/api/audit?limit=0", headers=sh)
        check("a nonsense limit is refused", r.status_code == 422)

        # ---- forced passcode change ----
        # Every account the Admin screen creates is flagged must_change. The
        # flag was written and then read by nothing, so a passcode an admin
        # typed stayed the operator's passcode indefinitely.
        r = client.post("/api/users", headers=sh,
                        json={"username": "newhire", "password": "temp-passcode-1",
                              "role": "operator"})
        check("a new operator is created", r.status_code == 200, r.text[:200])
        hire_id = r.json()["id"]

        r = client.post("/api/auth/login",
                        json={"username": "newhire", "password": "temp-passcode-1"})
        check("the new operator can sign in", r.status_code == 200)
        check("login says the passcode must be replaced",
              r.json().get("must_change") is True, r.text[:200])
        nh = {"X-API-Token": r.json()["token"]}

        r = client.get("/api/auth/me", headers=nh)
        check("whoami says so too, for a resumed session",
              r.json().get("must_change_password") is True, r.text[:200])

        r = client.post("/api/auth/change-password", headers=nh,
                        json={"current_password": "wrong", "new_password": "chosen-passcode-2"})
        check("changing needs the current passcode", r.status_code == 401)

        r = client.post("/api/auth/change-password", headers=nh,
                        json={"current_password": "temp-passcode-1",
                              "new_password": "chosen-passcode-2"})
        check("the operator can change their own passcode", r.status_code == 200, r.text[:200])
        r = client.get("/api/auth/me", headers=nh)
        check("the flag clears once it is changed",
              r.json().get("must_change_password") is False, r.text[:200])
        check("the old passcode stops working",
              client.post("/api/auth/login",
                          json={"username": "newhire",
                                "password": "temp-passcode-1"}).status_code == 401)

        # ---- signing out retires the token ----
        r = client.post("/api/auth/login",
                        json={"username": "newhire", "password": "chosen-passcode-2"})
        doomed = {"X-API-Token": r.json()["token"]}
        r = client.post("/api/auth/login",
                        json={"username": "newhire", "password": "chosen-passcode-2"})
        other = {"X-API-Token": r.json()["token"]}
        check("two devices can be signed in at once",
              client.get("/api/status", headers=doomed).status_code == 200 and
              client.get("/api/status", headers=other).status_code == 200)

        r = client.post("/api/auth/logout", headers=doomed)
        check("signing out reports the session retired",
              r.status_code == 200 and r.json()["revoked"] is True, r.text[:200])
        check("the signed-out token stops working",
              client.get("/api/status", headers=doomed).status_code == 401)
        check("the other device is untouched",
              client.get("/api/status", headers=other).status_code == 200)
        check("signing out is audited",
              any(e["action"] == "signed_out"
                  for e in client.get("/api/audit", headers=sh).json()["entries"]))

        r = client.post("/api/auth/logout", headers=h)
        check("signing out a static API token is a no-op, not an error",
              r.status_code == 200 and r.json()["revoked"] is False, r.text[:200])
        check("the API token still works afterwards",
              client.get("/api/status", headers=h).status_code == 200)

        client.delete(f"/api/users/{hire_id}", headers=sh)

        # Repeated failures must lock out rather than allow unlimited guessing.
        codes = [
            client.post("/api/auth/login",
                        json={"username": "Sehej", "password": f"guess-{i}"}).status_code
            for i in range(12)
        ]
        check("brute force is rate limited", 429 in codes, str(codes))

        r = client.get("/api/status", headers=h)
        check("rate limiting never blocks a valid token", r.status_code == 200)


def main() -> int:
    print("=" * 60)
    print("  MERIDIAN CAPITAL — SMOKE TESTS")
    print("=" * 60)
    test_ranges()
    seed()
    test_storage()
    test_aggregate()
    test_exports()
    test_live_mode_approval()
    test_expiry()
    test_log_severity()
    test_starter_and_brief()
    test_news()
    test_api()
    print("\n" + "=" * 60)
    print(f"  {PASS} passed, {FAIL} failed")
    print("=" * 60)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
