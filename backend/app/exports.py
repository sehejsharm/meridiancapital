"""Trade-result exports.

Any window the user asks for — a single day, a week, a month, a quarter, a
year, or an arbitrary range — resolved to a date pair and rendered as CSV,
JSON or a multi-sheet Excel workbook.
"""
from __future__ import annotations

import csv
import io
import json
from calendar import monthrange
from datetime import date, datetime, timedelta
from typing import Iterable, Literal, Optional

from . import db

Period = Literal["day", "week", "month", "quarter", "year", "all", "custom"]


# ------------------------------------------------------------------ ranges


def resolve_range(
    period: str,
    anchor: str | None = None,
    start: str | None = None,
    end: str | None = None,
) -> tuple[str, str, str]:
    """Return (start_date, end_date, human_label) for a period request.

    `anchor` is any date inside the desired window (defaults to today), so the
    phone can send "month" + "2026-08-04" and get the whole of August 2026.
    """
    if period == "custom":
        if not start or not end:
            raise ValueError("custom period requires start and end")
        s, e = _parse(start), _parse(end)
        if s > e:
            s, e = e, s
        return s.isoformat(), e.isoformat(), f"{s.isoformat()} to {e.isoformat()}"

    a = _parse(anchor) if anchor else date.today()

    if period == "day":
        return a.isoformat(), a.isoformat(), a.strftime("%d %b %Y")

    if period == "week":
        s = a - timedelta(days=a.weekday())          # Monday
        e = s + timedelta(days=6)
        return s.isoformat(), e.isoformat(), f"Week of {s.strftime('%d %b %Y')}"

    if period == "month":
        s = a.replace(day=1)
        e = a.replace(day=monthrange(a.year, a.month)[1])
        return s.isoformat(), e.isoformat(), a.strftime("%B %Y")

    if period == "quarter":
        q = (a.month - 1) // 3 + 1
        first_month = 3 * (q - 1) + 1
        s = date(a.year, first_month, 1)
        last_month = first_month + 2
        e = date(a.year, last_month, monthrange(a.year, last_month)[1])
        return s.isoformat(), e.isoformat(), f"Q{q} {a.year}"

    if period == "year":
        return date(a.year, 1, 1).isoformat(), date(a.year, 12, 31).isoformat(), str(a.year)

    if period == "all":
        rows = db.query("SELECT MIN(session_date) AS a, MAX(session_date) AS b FROM sessions")
        first = (rows[0]["a"] if rows and rows[0]["a"] else date.today().isoformat())
        last = (rows[0]["b"] if rows and rows[0]["b"] else date.today().isoformat())
        return first, last, "All time"

    raise ValueError(f"unknown period: {period}")


def _parse(raw: str) -> date:
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unparseable date: {raw}")


def filename(period: str, start: str, end: str, ext: str) -> str:
    stem = start if start == end else f"{start}_to_{end}"
    return f"meridian_{period}_{stem}.{ext}"


# ------------------------------------------------------------------ payload


def build_payload(start: str, end: str, label: str, include_events: bool = False,
                  slot: Optional[int] = None) -> dict:
    """`slot=None` reports the whole book; a slot reports one algorithm.

    Without the log this is the trade record: what was bought and sold, when,
    at what price, for what, and the risk-adjusted view of the result. That is
    the document you send to someone.

    With it, everything the algorithm said is attached as well — every printed
    line and every structured event, in order, plus a mark for every minute it
    was running. That is the document you read when you want to know why it
    did something, and it is much larger.
    """
    trades = db.trades_between(start, end, slot=slot)
    sessions = db.sessions_between(start, end, slot=slot)
    summary = db.aggregate(start, end, slot=slot)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "slot": slot,
        "range": {"start": start, "end": end, "label": label},
        "summary": summary,
        "sessions": sessions,
        "trades": trades,
    }
    if include_events:
        # `kind != 'log'` used to be in this query, which quietly dropped the
        # algorithm's own printed output — the per-minute narration that is the
        # main reason anyone asks for the log at all. Everything is included
        # now, in the order it happened.
        sql = ("""SELECT ts, session_date, slot, kind, level, message, payload
                    FROM events
                   WHERE session_date >= ? AND session_date <= ?""")
        params: list = [start, end]
        if slot is not None:
            sql += " AND slot = ?"
            params.append(slot)
        payload["events"] = db.query(sql + " ORDER BY session_date, ts, id", params)
        payload["minutes"] = db.equity_marks_between(start, end, slot=slot)
    return payload


# ------------------------------------------------------------------ writers


TRADE_HEADERS = [
    ("session_date", "Date"), ("mode", "Mode"), ("entry_time", "Entry Time"),
    ("exit_time", "Exit Time"), ("hold_min", "Hold (min)"), ("symbol", "Symbol"),
    ("opt_type", "Type"), ("strike", "Strike"), ("qty", "Qty"),
    ("avg_entry", "Entry Price"), ("exit_fill", "Exit Price"),
    ("gross_pnl", "Gross P&L"), ("brokerage", "Brokerage"), ("stt", "STT"),
    ("exch_txn", "Exchange"), ("sebi", "SEBI"), ("stamp", "Stamp"),
    ("gst", "GST"), ("charges", "Total Charges"), ("net_pnl", "Net P&L"),
    ("reason", "Exit Reason"), ("stage", "Ladder Stage"),
    ("entry_reason", "Entry Reason"), ("spot_at_entry", "NIFTY at Entry"),
    ("garch_vol", "GARCH"), ("entry_iv", "Entry IV"),
    ("model_prem", "BSM Model Premium"), ("lot_cost", "Lot Cost"),
    ("real_margin", "Margin"), ("risk_rs", "Risk (Rs)"),
    ("day_pnl", "Day P&L"), ("equity_after", "Equity After"),
    ("latency_ms", "Avg Latency (ms)"),
]

# Risk-adjusted return. Computed in db.aggregate() and, until now, shown only
# on screen — which meant the report you actually send to someone had the P&L
# and none of the context for judging whether it was worth the risk taken.
RISK_ROWS = [
    ("sharpe", "Sharpe ratio", "annualised, from daily returns, rf 6.5%"),
    ("sortino", "Sortino ratio", "downside deviation only"),
    ("calmar", "Calmar ratio", "return over max drawdown, 90+ days"),
    ("expectancy", "Expectancy per trade", "what one more trade is worth"),
    ("avg_win", "Average win", ""),
    ("avg_loss", "Average loss", ""),
    ("win_loss_ratio", "Win/loss ratio", "average win over average loss"),
    ("daily_vol_pct", "Daily volatility %", "standard deviation of daily returns"),
    ("best_day", "Best day", ""),
    ("worst_day", "Worst day", ""),
    ("trading_period_days", "Period length (days)", ""),
]

MINUTE_HEADERS = [
    ("ts", "Timestamp"), ("session_date", "Date"), ("slot", "Slot"),
    ("equity", "Equity"), ("day_pnl", "Day P&L"),
    ("open_position", "In a position"), ("unrealised", "Unrealised"),
]

EVENT_HEADERS = [
    ("ts", "Timestamp"), ("session_date", "Date"), ("slot", "Slot"),
    ("kind", "Kind"), ("level", "Level"), ("message", "Message"),
]

SESSION_HEADERS = [
    ("session_date", "Date"), ("mode", "Mode"), ("trades", "Trades"),
    ("open_equity", "Open Equity"), ("close_equity", "Close Equity"),
    ("day_pnl", "Day P&L"), ("charges", "Charges"), ("win_rate", "Win Rate %"),
    ("profit_factor", "Profit Factor"), ("killed", "Kill Switch"),
    ("chop_blocked", "Chop Blocked"), ("chop_score", "Chop Score"),
    ("garch", "GARCH %"), ("adx", "ADX"), ("vol_regime", "Vol Regime"),
    ("trend", "Trend"), ("direction", "Direction"), ("efficiency", "Efficiency"),
    ("day_range_pts", "Day Range (pts)"), ("peak_equity", "Peak Equity"),
    ("drawdown_pct", "Drawdown %"), ("avg_latency_ms", "Avg Latency (ms)"),
]


def to_csv(payload: dict) -> str:
    """One readable document: summary block, then sessions, then trades."""
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    s = payload["summary"]
    rng = payload["range"]

    w.writerow(["MERIDIAN CAPITAL — TRADE RESULTS"])
    w.writerow(["Period", rng["label"]])
    w.writerow(["From", rng["start"], "To", rng["end"]])
    w.writerow(["Generated", payload["generated_at"]])
    w.writerow([])

    w.writerow(["SUMMARY"])
    for key, label in [
        ("sessions", "Sessions"), ("trading_days", "Days with trades"),
        ("trades", "Total trades"), ("wins", "Wins"), ("losses", "Losses"),
        ("win_rate", "Win rate %"), ("profit_factor", "Profit factor"),
        ("gross_profit", "Gross profit"), ("gross_loss", "Gross loss"),
        ("net_pnl", "Net P&L"), ("charges", "Total charges"),
        ("avg_trade", "Average trade"), ("best_trade", "Best trade"),
        ("worst_trade", "Worst trade"), ("avg_hold_min", "Avg hold (min)"),
        ("open_equity", "Opening equity"), ("close_equity", "Closing equity"),
        ("return_pct", "Return %"), ("peak_equity", "Peak equity"),
        ("max_drawdown_pct", "Max drawdown %"),
        ("days_blocked_chop", "Days blocked by chop filter"),
        ("days_killed", "Days kill switch hit"),
    ]:
        w.writerow([label, _fmt(s.get(key))])
    w.writerow([])

    w.writerow(["RISK AND RETURN"])
    for key, label, note in RISK_ROWS:
        w.writerow([label, _fmt(s.get(key)), note])
    w.writerow([])

    w.writerow(["DAILY SESSIONS"])
    w.writerow([label for _, label in SESSION_HEADERS])
    for row in payload["sessions"]:
        w.writerow([_fmt(row.get(k)) for k, _ in SESSION_HEADERS])
    w.writerow([])

    w.writerow(["TRADES"])
    w.writerow([label for _, label in TRADE_HEADERS])
    for row in payload["trades"]:
        w.writerow([_fmt(row.get(k)) for k, _ in TRADE_HEADERS])

    if payload.get("minutes"):
        w.writerow([])
        w.writerow(["MINUTE LOG", f"{len(payload['minutes'])} marks"])
        w.writerow([label for _, label in MINUTE_HEADERS])
        for m in payload["minutes"]:
            w.writerow([_fmt(m.get(k)) for k, _ in MINUTE_HEADERS])

    if payload.get("events"):
        w.writerow([])
        w.writerow(["ACTIVITY LOG", f"{len(payload['events'])} entries"])
        w.writerow([label for _, label in EVENT_HEADERS])
        for e in payload["events"]:
            w.writerow([_fmt(e.get(k)) for k, _ in EVENT_HEADERS])

    return buf.getvalue()


def to_json(payload: dict) -> str:
    return json.dumps(payload, indent=2, default=str)


def to_xlsx(payload: dict) -> bytes:
    """Multi-sheet workbook — the one people actually open on a laptop later."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="0B1220")
    head_font = Font(color="E8EDF7", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    money = '#,##0.00;[Red]-#,##0.00'

    def _autosize(ws, widths: Iterable[int]) -> None:
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "MERIDIAN CAPITAL — TRADE RESULTS"
    ws["A1"].font = title_font
    ws["A2"] = "Period"
    ws["B2"] = payload["range"]["label"]
    ws["A3"] = "From"
    ws["B3"] = payload["range"]["start"]
    ws["C3"] = "To"
    ws["D3"] = payload["range"]["end"]
    ws["A4"] = "Generated"
    ws["B4"] = payload["generated_at"]

    s = payload["summary"]
    rows = [
        ("Sessions", s.get("sessions")), ("Days with trades", s.get("trading_days")),
        ("Total trades", s.get("trades")), ("Wins", s.get("wins")),
        ("Losses", s.get("losses")), ("Win rate %", s.get("win_rate")),
        ("Profit factor", s.get("profit_factor")),
        ("Gross profit", s.get("gross_profit")), ("Gross loss", s.get("gross_loss")),
        ("Net P&L", s.get("net_pnl")), ("Total charges", s.get("charges")),
        ("Average trade", s.get("avg_trade")), ("Best trade", s.get("best_trade")),
        ("Worst trade", s.get("worst_trade")), ("Avg hold (min)", s.get("avg_hold_min")),
        ("Opening equity", s.get("open_equity")), ("Closing equity", s.get("close_equity")),
        ("Return %", s.get("return_pct")), ("Peak equity", s.get("peak_equity")),
        ("Max drawdown %", s.get("max_drawdown_pct")),
        ("Days blocked by chop filter", s.get("days_blocked_chop")),
        ("Days kill switch hit", s.get("days_killed")),
    ]
    ws["A6"] = "SUMMARY"
    ws["A6"].font = Font(bold=True)
    for i, (label, value) in enumerate(rows, start=7):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)

    risk_at = 7 + len(rows) + 1
    ws.cell(row=risk_at, column=1, value="RISK AND RETURN").font = Font(bold=True)
    for i, (key, label, note) in enumerate(RISK_ROWS, start=risk_at + 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=s.get(key))
        ws.cell(row=i, column=3, value=note)
    _autosize(ws, [30, 20, 44, 20])

    def _sheet(name: str, headers: list[tuple[str, str]], data: list[dict]) -> None:
        sh = wb.create_sheet(name)
        for col, (_, label) in enumerate(headers, start=1):
            c = sh.cell(row=1, column=col, value=label)
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        for r, row in enumerate(data, start=2):
            for col, (key, _) in enumerate(headers, start=1):
                cell = sh.cell(row=r, column=col, value=row.get(key))
                if key in ("gross_pnl", "net_pnl", "charges", "avg_entry", "exit_fill",
                           "open_equity", "close_equity", "day_pnl", "equity_after",
                           "peak_equity", "risk_rs", "lot_cost", "real_margin"):
                    cell.number_format = money
        sh.freeze_panes = "A2"
        # A message column sized to the word "Message" is unreadable, and a
        # timestamp column sized to "Timestamp" clips to ####.
        wide = {"message": 100, "ts": 24, "entry_reason": 46, "reason": 22}
        _autosize(sh, [wide.get(key, max(12, min(28, len(label) + 4)))
                       for key, label in headers])

    _sheet("Trades", TRADE_HEADERS, payload["trades"])
    _sheet("Daily Sessions", SESSION_HEADERS, payload["sessions"])

    if payload.get("minutes"):
        _sheet("Minute Log", MINUTE_HEADERS, payload["minutes"])

    if payload.get("events"):
        # Excel refuses a sheet past 1,048,576 rows, and a month of a chatty
        # algorithm can approach that. Spilling into numbered sheets keeps the
        # whole log rather than truncating it without saying so.
        LIMIT = 1_000_000
        chunks = [payload["events"][i:i + LIMIT]
                  for i in range(0, len(payload["events"]), LIMIT)] or [[]]
        for n, chunk in enumerate(chunks, start=1):
            name = "Activity Log" if len(chunks) == 1 else f"Activity Log {n}"
            _sheet(name, EVENT_HEADERS, chunk)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _fmt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:.2f}"
    return str(v)


def _xml(text: str) -> str:
    """Escape for reportlab's paragraph markup parser."""
    return (str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def to_pdf(payload: dict) -> bytes:
    """A report someone can send to an accountant without reformatting it.

    reportlab is imported here rather than at module scope: the PDF path is
    occasional, and the trading server runs on 1 GB, so the ~20 MB the library
    pulls in should not be resident for every process that touches exports.
    """
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    INK = colors.HexColor("#101014")
    GOLD = colors.HexColor("#8a6d1f")
    RULE = colors.HexColor("#d8d2c4")
    UP = colors.HexColor("#0f7a4a")
    DOWN = colors.HexColor("#b3242f")

    rng = payload.get("range") or {}
    summary = payload.get("summary") or {}
    trades = payload.get("trades") or []
    sessions = payload.get("sessions") or []

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=14 * mm,
        title=f"Meridian Capital — {rng.get('label', '')}",
        author="Meridian Capital",
    )

    ss = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=ss["Title"], fontName="Helvetica-Bold",
                        fontSize=17, textColor=INK, alignment=0, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontName="Helvetica",
                         fontSize=9, textColor=colors.HexColor("#6a6558"))
    h2 = ParagraphStyle("h2", parent=ss["Heading2"], fontName="Helvetica-Bold",
                        fontSize=11, textColor=GOLD, spaceBefore=10, spaceAfter=5)
    small = ParagraphStyle("small", parent=ss["Normal"], fontName="Helvetica",
                           fontSize=7.5, textColor=INK)

    story = [
        Paragraph("MERIDIAN CAPITAL", h1),
        Paragraph(
            f"{rng.get('label', '')} &nbsp;·&nbsp; {rng.get('start', '')} to "
            f"{rng.get('end', '')} &nbsp;·&nbsp; generated "
            f"{payload.get('generated_at', '')} IST", sub),
        Spacer(1, 7),
    ]

    def money(v) -> str:
        try:
            return f"{float(v):,.2f}"
        except (TypeError, ValueError):
            return "—"

    net = summary.get("net_pnl") or 0
    cards = [
        ("Net P&L", money(net), UP if net >= 0 else DOWN),
        ("Trades", str(summary.get("trades") or 0), INK),
        ("Win rate", f"{summary.get('win_rate') or 0:.1f}%", INK),
        ("Profit factor",
         "—" if summary.get("profit_factor") in (None, "") else f"{summary['profit_factor']:.2f}",
         INK),
        ("Charges", money(summary.get("charges")), INK),
        ("Return", f"{summary.get('return_pct') or 0:.2f}%", INK),
        ("Max drawdown", f"{summary.get('max_drawdown_pct') or 0:.2f}%", DOWN),
    ]
    head = Table(
        [[Paragraph(f"<font size=7 color='#6a6558'>{k.upper()}</font>", small) for k, _, _ in cards],
         [Paragraph(f"<font size=13 color='#{c.hexval()[2:]}'><b>{v}</b></font>", small)
          for _, v, c in cards]],
        colWidths=[doc.width / len(cards)] * len(cards),
    )
    head.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, RULE),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, RULE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#faf8f3")),
    ]))
    story += [head, Spacer(1, 4)]

    def table(title: str, headers: list[tuple[str, str]], rows: list[dict],
              highlight: str = "") -> None:
        story.append(Paragraph(title, h2))
        if not rows:
            story.append(Paragraph("<i>Nothing in this window.</i>", small))
            return
        keys = [k for k, _ in headers]
        data = [[Paragraph(f"<b>{lbl}</b>", small) for _, lbl in headers]]
        for r in rows:
            # Cell text is parsed as markup, so an exit reason containing "&"
            # or "<" would abort the whole render. Entry reasons are free text
            # written by the algorithm, so this is not hypothetical.
            data.append([Paragraph(_xml(_fmt(r.get(k))), small) for k in keys])

        t = Table(data, repeatRows=1, colWidths=[doc.width / len(keys)] * len(keys))
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2eee4")),
            ("GRID", (0, 0), (-1, -1), 0.3, RULE),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#fbfaf7")]),
        ]
        # Colour the P&L column so a losing row is visible at a glance.
        if highlight in keys:
            col = keys.index(highlight)
            for i, r in enumerate(rows, start=1):
                try:
                    val = float(r.get(highlight) or 0)
                except (TypeError, ValueError):
                    continue
                style.append(("TEXTCOLOR", (col, i), (col, i), UP if val >= 0 else DOWN))
        t.setStyle(TableStyle(style))
        story.append(t)

    # The full trade table is far too wide for a page, so the PDF carries the
    # columns a human reads and the CSV/XLSX keep every field.
    pdf_trade_cols = [
        ("session_date", "Date"), ("entry_time", "Entry"), ("exit_time", "Exit"),
        ("symbol", "Symbol"), ("qty", "Qty"), ("avg_entry", "Entry"),
        ("exit_fill", "Exit"), ("charges", "Charges"), ("net_pnl", "Net P&L"),
        ("reason", "Exit Reason"), ("stage", "Stage"),
    ]
    pdf_session_cols = [
        ("session_date", "Date"), ("trades", "Trades"), ("open_equity", "Open"),
        ("close_equity", "Close"), ("day_pnl", "Day P&L"), ("charges", "Charges"),
        ("win_rate", "Win %"), ("drawdown_pct", "DD %"), ("trend", "Trend"),
    ]

    # Risk-adjusted return, laid out as pairs so it fits beside the tables
    # rather than running down a whole page.
    def risk_block() -> None:
        present = [(lbl, summary.get(key), note) for key, lbl, note in RISK_ROWS
                   if summary.get(key) is not None]
        if not present:
            return
        story.append(Paragraph("Risk and return", h2))
        cells = []
        for lbl, val, note in present:
            shown = f"{val:,.2f}" if isinstance(val, (int, float)) else str(val)
            cells.append(Paragraph(
                f"<font size=6.5 color='#6a6558'>{_xml(lbl).upper()}</font><br/>"
                f"<font size=11><b>{shown}</b></font>"
                + (f"<br/><font size=6 color='#8a857a'>{_xml(note)}</font>" if note else ""),
                small))
        per_row = 4
        grid = [cells[i:i + per_row] for i in range(0, len(cells), per_row)]
        grid[-1] += [Paragraph("", small)] * (per_row - len(grid[-1]))
        t = Table(grid, colWidths=[doc.width / per_row] * per_row)
        t.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, RULE),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, RULE),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(t)

    risk_block()
    table(f"Daily sessions ({len(sessions)})", pdf_session_cols, sessions,
          highlight="day_pnl")
    if trades:
        story.append(PageBreak())
    table(f"Trades ({len(trades)})", pdf_trade_cols, trades, highlight="net_pnl")

    # ---- the activity log ----
    #
    # A chatty algorithm produces tens of thousands of lines a day, and a
    # month of them is not a document anybody can open. The PDF carries as
    # much as stays usable and says plainly where the rest is, rather than
    # truncating silently or refusing to render at all.
    minutes = payload.get("minutes") or []
    if minutes:
        story.append(PageBreak())
        table(f"Minute log ({len(minutes):,} marks)",
              [("ts", "Time"), ("session_date", "Date"), ("slot", "Slot"),
               ("equity", "Equity"), ("day_pnl", "Day P&L"),
               ("open_position", "In a position"), ("unrealised", "Unrealised")],
              minutes, highlight="day_pnl")

    events = payload.get("events") or []
    if events:
        PDF_LOG_LIMIT = 4000
        shown = events[:PDF_LOG_LIMIT]
        story.append(PageBreak())
        story.append(Paragraph(f"Activity log ({len(events):,} entries)", h2))
        if len(events) > PDF_LOG_LIMIT:
            story.append(Paragraph(
                f"<i>The first {PDF_LOG_LIMIT:,} are printed here. The CSV and "
                f"Excel exports of this same range carry all {len(events):,}.</i>",
                small))
            story.append(Spacer(1, 4))

        log_style = ParagraphStyle("log", parent=small, fontName="Courier",
                                   fontSize=6.2, leading=7.6)
        head = [Paragraph(f"<b>{h}</b>", small)
                for h in ("Time", "Slot", "Kind", "Level", "Message")]
        widths = [doc.width * w for w in (0.13, 0.05, 0.10, 0.07, 0.65)]

        # Built in blocks: one Table of 40,000 cells is slow to lay out and
        # holds the whole thing in memory, which the trading box cannot spare.
        BLOCK = 500
        for i in range(0, len(shown), BLOCK):
            data = [head] if i == 0 else [head]
            for e in shown[i:i + BLOCK]:
                data.append([
                    Paragraph(_xml(str(e.get("ts") or "")[11:23]), log_style),
                    Paragraph(_xml(_fmt(e.get("slot"))), log_style),
                    Paragraph(_xml(e.get("kind") or ""), log_style),
                    Paragraph(_xml(e.get("level") or ""), log_style),
                    Paragraph(_xml(e.get("message") or ""), log_style),
                ])
            t = Table(data, repeatRows=1, colWidths=widths)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2eee4")),
                ("GRID", (0, 0), (-1, -1), 0.25, RULE),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 1.2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2),
            ]))
            story.append(t)

    def chrome(canvas, doc_):
        canvas.saveState()
        canvas.setStrokeColor(RULE)
        canvas.setLineWidth(0.5)
        canvas.line(doc_.leftMargin, 9 * mm,
                    doc_.pagesize[0] - doc_.rightMargin, 9 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#8a857a"))
        canvas.drawString(doc_.leftMargin, 5.5 * mm,
                          "Meridian Capital — algorithmic trading record")
        canvas.drawRightString(doc_.pagesize[0] - doc_.rightMargin, 5.5 * mm,
                               f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=chrome, onLaterPages=chrome)
    return buf.getvalue()
